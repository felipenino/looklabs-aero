#!/usr/bin/env python3
"""
Gera os dados reais do LookLab a partir de:
  - Planilha Looklab + Aero - v1 Piloto-15-junho.xlsx  (info das peças)
  - CSV-APROVADOS/looklab_aeropostale-parte1.csv + parte2.csv  (looks)

Saídas:
  - public/images/pecas-real/<SKU>.png   (imagens copiadas das peças usadas)
  - src/lib/real-data.ts                  (dados consumidos pelo app)

Regras:
  - SKU sem arquivo exato -> usa 1ª variante de cor (<SKU>-*.png)  [resolver]
  - Looks com alguma peça irrecuperável são descartados
  - Slot definido pela categoria (não pela ordem do CSV):
      base (CAMISETA/POLO) -> slot 1
      camada (MOLETOM/JAQUETA, e CAMISA quando há base) -> slot 2
      calçado (CALÇADOS) -> slot 3
      baixo (CALÇA/BERMUDA) -> slot 4
      CAMISA vira base (slot 1) quando não há CAMISETA/POLO no look
"""
import csv, os, shutil, json, sys, hashlib

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "Planilha Looklab + Aero - v1 Piloto-15-junho.xlsx")
CSV_DIR = os.path.join(ROOT, "CSV-APROVADOS")
PECAS_DIR = os.path.join(ROOT, "pecas-final-piloto")
OUT_IMG = os.path.join(ROOT, "public", "images", "pecas-real")
OUT_TS = os.path.join(ROOT, "src", "lib", "real-data.ts")

MASC_PERSONA = "a1b2c3d4-0001-4000-8000-000000000001"

# classificacao (planilha) -> category id (categories.ts)
CAT_ID = {
    "CAMISETA": "b2c3d4e5-0001-4000-8000-000000000001",  # Camiseta
    "POLO":     "b2c3d4e5-0001-4000-8000-000000000002",  # Top / Blusa
    "CAMISA":   "b2c3d4e5-0001-4000-8000-000000000002",  # Top / Blusa
    "CALÇA":    "b2c3d4e5-0001-4000-8000-000000000003",  # Calça / Jeans
    "BERMUDA":  "b2c3d4e5-0001-4000-8000-000000000004",  # Bermuda
    "JAQUETA":  "b2c3d4e5-0001-4000-8000-000000000009",  # Moletom / Jaqueta (unificado)
    "MOLETOM":  "b2c3d4e5-0001-4000-8000-000000000009",  # Moletom / Jaqueta
    "CALÇADOS": "b2c3d4e5-0001-4000-8000-000000000010",  # Acessório
}

BASE = {"CAMISETA", "POLO"}
OUTER = {"MOLETOM", "JAQUETA"}
BOTTOM = {"CALÇA", "BERMUDA"}
SHOES = {"CALÇADOS"}


def norm(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def load_pieces():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    info = {}
    for sh in ["MASCULINO"]:
        w = wb[sh]
        for r in range(2, w.max_row + 1):
            sku = norm(w.cell(r, 2).value)
            if sku and sku not in info:
                info[sku] = {
                    "cor": w.cell(r, 5).value,
                    "pdv": w.cell(r, 6).value,
                    "nome": w.cell(r, 7).value,
                    "classif": w.cell(r, 10).value,
                }
    return info


def build_resolver():
    files = [os.path.splitext(f)[0] for f in os.listdir(PECAS_DIR) if f.lower().endswith(".png")]
    real_name = {f.lower(): f for f in files}  # lower -> real filename (case)

    def resolve(sku):
        if sku.lower() in real_name:
            return real_name[sku.lower()]
        cands = [f for f in files if f.lower().startswith(sku.lower() + "-")]
        return cands[0] if cands else None

    return resolve


def slot_for(classif, has_base):
    if classif in BASE:
        return 1
    if classif in OUTER:
        return 2
    if classif == "CAMISA":
        return 2 if has_base else 1
    if classif in BOTTOM:
        return 4
    if classif in SHOES:
        return 3
    return None


def main():
    info = load_pieces()
    resolve = build_resolver()

    used_skus = {}      # sku -> resolved filename
    looks = []          # {id, anchor, name, tags, pieces:[{slot,sku}]}
    dropped = 0

    for part in ["parte1", "parte2"]:
        path = os.path.join(CSV_DIR, f"looklab_aeropostale-{part}.csv")
        with open(path) as f:
            for row in csv.DictReader(f):
                pecas = [p for p in row["pecas"].split("|") if p]
                # resolve all images first
                resolved = {p: resolve(p) for p in pecas}
                if any(v is None for v in resolved.values()):
                    dropped += 1
                    continue
                classifs = {p: info.get(p, {}).get("classif") for p in pecas}
                has_base = any(classifs[p] in BASE for p in pecas)
                slots = {}
                collision = False
                for p in pecas:
                    s = slot_for(classifs[p], has_base)
                    if s is None or s in slots:
                        collision = True
                        break
                    slots[s] = p
                if collision:
                    dropped += 1
                    continue
                for p in pecas:
                    used_skus[p] = resolved[p]
                looks.append({
                    "id": f"{part[-1]}-{row['look_id']}",
                    "anchor": row["anchor"],
                    "name": row["nome"],
                    "tags": row["tags"],
                    "pieces": [{"slot": s, "sku": sku} for s, sku in sorted(slots.items())],
                })

    # copy images com hash de conteúdo no nome (cache-busting sem query string)
    os.makedirs(OUT_IMG, exist_ok=True)
    img_name = {}  # sku -> nome final do arquivo (com hash)
    for sku, fname in used_skus.items():
        src = os.path.join(PECAS_DIR, fname + ".png")
        with open(src, "rb") as fh:
            ver = hashlib.md5(fh.read()).hexdigest()[:8]
        out = f"{sku}.{ver}.png"
        img_name[sku] = out
        dst = os.path.join(OUT_IMG, out)
        if not os.path.exists(dst):
            shutil.copyfile(src, dst)
    # remove variantes antigas (mesmo sku, hash diferente)
    keep = set(img_name.values())
    for f in os.listdir(OUT_IMG):
        if f.endswith(".png") and f not in keep:
            os.remove(os.path.join(OUT_IMG, f))

    # pieces payload (only used pieces)
    pieces_payload = {}
    for sku in used_skus:
        meta = info.get(sku, {})
        classif = meta.get("classif")
        pdv = meta.get("pdv")
        pieces_payload[sku] = {
            "id": sku,
            "persona_id": MASC_PERSONA,
            "category_id": CAT_ID.get(classif, ""),
            "image_url": f"/images/pecas-real/{img_name[sku]}",
            "name": meta.get("nome") or sku,
            "cor": meta.get("cor"),
            "preco_pdv": float(pdv) if isinstance(pdv, (int, float)) else None,
            "classif": classif,
        }

    # anchors = pieces that are anchor of at least one look
    anchors = sorted({lk["anchor"] for lk in looks})

    emit_ts(pieces_payload, looks, anchors)
    print(f"Looks gerados: {len(looks)}  |  descartados: {dropped}")
    print(f"Peças usadas: {len(used_skus)}  |  âncoras: {len(anchors)}")
    print(f"Imagens em: {OUT_IMG}")
    print(f"Dados em:   {OUT_TS}")


def emit_ts(pieces, looks, anchors):
    def js(v):
        return json.dumps(v, ensure_ascii=False)

    lines = []
    lines.append("// AUTO-GERADO por scripts/gen_real_data.py — NÃO editar à mão.")
    lines.append("// Fonte: Planilha v1 Piloto-15-junho.xlsx + CSV-APROVADOS/*.csv")
    lines.append('import type { Piece, LookDetailFull } from "@/types/database";')
    lines.append("")
    lines.append("interface RealPiece extends Piece { classif: string | null; }")
    lines.append("")
    # pieces map
    lines.append("const PIECES: Record<string, RealPiece> = {")
    for sku, p in pieces.items():
        lines.append(f"  {js(sku)}: {{")
        lines.append(f"    id: {js(p['id'])}, persona_id: {js(p['persona_id'])}, category_id: {js(p['category_id'])},")
        lines.append(f"    image_url: {js(p['image_url'])}, name: {js(p['name'])}, sku: {js(sku)},")
        lines.append(f"    is_active: true, cor: {js(p['cor'])}, preco_pdv: {js(p['preco_pdv'])},")
        lines.append(f"    tamanho: null, grade: null, subcategoria: null, codigo_aero: {js(sku)},")
        lines.append(f"    quantidade: null, status_looklab: \"Ativo\", classif: {js(p['classif'])},")
        lines.append("  },")
    lines.append("};")
    lines.append("")
    # looks
    lines.append("interface RealLook { id: string; anchor: string; name: string; tags: string; pieces: { slot: number; sku: string }[]; }")
    lines.append("const LOOKS: RealLook[] = [")
    for lk in looks:
        pcs = ", ".join(f"{{ slot: {pc['slot']}, sku: {js(pc['sku'])} }}" for pc in lk["pieces"])
        lines.append(f"  {{ id: {js(lk['id'])}, anchor: {js(lk['anchor'])}, name: {js(lk['name'])}, tags: {js(lk['tags'])}, pieces: [{pcs}] }},")
    lines.append("];")
    lines.append("")
    lines.append(f"const ANCHORS: string[] = {js(anchors)};")
    lines.append("")
    lines.append("""// ── API (mesma assinatura do mock antigo) ──

/** Peças-âncora de uma persona+categoria (o que aparece na grade). */
export function getRealPieces(personaId: string, categoryId: string): Piece[] {
  return ANCHORS
    .map((sku) => PIECES[sku])
    .filter((p) => p && p.persona_id === personaId && p.category_id === categoryId);
}

/** Categorias (ids) que possuem peças-âncora para a persona. */
export function getRealCategoryIds(personaId: string): string[] {
  const ids = new Set<string>();
  for (const sku of ANCHORS) {
    const p = PIECES[sku];
    if (p && p.persona_id === personaId) ids.add(p.category_id);
  }
  return [...ids];
}

/** Looks reais em que a peça é âncora. */
export function getRealLooksForPiece(pieceId: string): LookDetailFull[] {
  return LOOKS.filter((lk) => lk.anchor === pieceId).map((lk) => ({
    id: lk.id,
    name: lk.name,
    pieces: lk.pieces
      .map((pc) => {
        const p = PIECES[pc.sku];
        if (!p) return null;
        return { slot: pc.slot, imageUrl: p.image_url, name: p.name, categoryName: p.classif };
      })
      .filter(Boolean) as LookDetailFull["pieces"],
  }));
}
""")
    with open(OUT_TS, "w") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    main()
